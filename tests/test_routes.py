from io import BytesIO

from openpyxl import load_workbook

from database import Inspection, Project, db


def test_index_route_serves_html(client):
    response = client.get("/")

    assert response.status_code == 200
    assert b"Machine Inspection" in response.data


def test_save_get_delete_history_routes(client, app, image_bytes_factory):
    response = client.post(
        "/api/save-inspection",
        data={
            "machine": "Route Machine",
            "issue": "Route issue",
            "remark": "Route note",
            "images": (image_bytes_factory(), "route.png"),
        },
        content_type="multipart/form-data",
    )
    payload = response.get_json()

    assert response.status_code == 201
    assert payload["success"] is True
    inspection_id = payload["id"]

    history_response = client.get("/api/get-history?machine=route%20machine")
    history_payload = history_response.get_json()
    assert history_response.status_code == 200
    assert [item["id"] for item in history_payload["data"]] == [inspection_id]

    delete_response = client.delete(f"/api/delete/{inspection_id}")
    assert delete_response.status_code == 200
    assert delete_response.get_json()["success"] is True
    with app.app_context():
        assert db.session.get(Inspection, inspection_id) is None


def test_export_excel_route_returns_workbook(client):
    client.post(
        "/api/save-inspection",
        data={
            "machine": "Export Route Machine",
            "issue": "Export route issue",
        },
        content_type="multipart/form-data",
    )

    response = client.get("/api/export-excel")

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(response.data))
    assert workbook.active.title == "Inspection History"


def test_project_routes_and_history_scope(client):
    projects_response = client.get("/api/projects")
    projects_payload = projects_response.get_json()
    default_project_id = projects_payload["current_project_id"]

    create_response = client.post("/api/projects", json={"name": "Route Project"})
    create_payload = create_response.get_json()
    route_project_id = create_payload["project"]["id"]

    assert create_response.status_code == 201

    client.post(
        "/api/save-inspection",
        data={
            "project_id": route_project_id,
            "machine": "Route Project Machine",
            "issue": "Project issue",
        },
        content_type="multipart/form-data",
    )

    default_history = client.get(f"/api/get-history?project_id={default_project_id}").get_json()["data"]
    project_history = client.get(f"/api/get-history?project_id={route_project_id}").get_json()["data"]

    assert default_history == []
    assert len(project_history) == 1
    assert project_history[0]["project_id"] == route_project_id

    delete_response = client.delete(f"/api/projects/{route_project_id}")
    assert delete_response.status_code == 400

    with client.application.app_context():
        assert db.session.get(Project, route_project_id) is not None


def test_delete_empty_project_route_succeeds(client):
    create_response = client.post("/api/projects", json={"name": "Empty Route Project"})
    project_id = create_response.get_json()["project"]["id"]

    delete_response = client.delete(f"/api/projects/{project_id}")
    payload = delete_response.get_json()

    assert delete_response.status_code == 200
    assert payload["success"] is True
    with client.application.app_context():
        assert db.session.get(Project, project_id) is None


def test_rename_project_route_updates_project(client):
    create_response = client.post("/api/projects", json={"name": "Old Route Project"})
    project_id = create_response.get_json()["project"]["id"]

    rename_response = client.put(f"/api/projects/{project_id}", json={"name": "New Route Project"})
    payload = rename_response.get_json()

    assert rename_response.status_code == 200
    assert payload["success"] is True
    assert payload["project"]["name"] == "New Route Project"

    projects = client.get("/api/projects").get_json()["data"]
    assert "New Route Project" in [project["name"] for project in projects]


def test_rename_project_route_rejects_duplicate_name(client):
    first_response = client.post("/api/projects", json={"name": "First Route Project"})
    first_project_id = first_response.get_json()["project"]["id"]
    client.post("/api/projects", json={"name": "Second Route Project"})

    rename_response = client.put(f"/api/projects/{first_project_id}", json={"name": "second route project"})
    payload = rename_response.get_json()

    assert rename_response.status_code == 409
    assert payload["success"] is False


def test_export_excel_route_is_scoped_by_project(client):
    projects_payload = client.get("/api/projects").get_json()
    default_project_id = projects_payload["current_project_id"]
    create_response = client.post("/api/projects", json={"name": "Export Scope Project"})
    project_id = create_response.get_json()["project"]["id"]

    client.post(
        "/api/save-inspection",
        data={
            "project_id": default_project_id,
            "machine": "Default Export Machine",
            "issue": "Default export issue",
        },
        content_type="multipart/form-data",
    )
    client.post(
        "/api/save-inspection",
        data={
            "project_id": project_id,
            "machine": "Scoped Export Machine",
            "issue": "Scoped export issue",
        },
        content_type="multipart/form-data",
    )

    response = client.get(f"/api/export-excel?project_id={project_id}")
    workbook = load_workbook(BytesIO(response.data))
    worksheet = workbook.active

    assert response.status_code == 200
    assert worksheet.cell(row=2, column=1).value == "Scoped Export Machine"
    assert worksheet.max_row == 2
