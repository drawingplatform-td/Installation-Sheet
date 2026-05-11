import json
from datetime import datetime
from types import SimpleNamespace

from openpyxl import load_workbook
from PIL import Image

from inspection_app.services import export_service
from inspection_app.utils.inspection_utils import SEVERITY_RANK_MAP
from inspection_app.utils.workbook_utils import create_excel_image_buffer


HIGH = {rank: value for value, rank in SEVERITY_RANK_MAP.items()}[3]


def make_record(machine="Machine A", image_links=None):
    return SimpleNamespace(
        machine=machine,
        image_links=json.dumps(image_links or []),
        issue="Issue text",
        severity=HIGH,
        remark="Remark text",
        timestamp=datetime(2026, 1, 1, 8, 30, 0),
    )


def create_export_images(folder, count):
    links = []
    for index in range(count):
        filename = f"export-{index}.png"
        Image.new("RGB", (80, 60), (index * 30, 80, 160)).save(folder / filename)
        links.append(f"/uploads/{filename}")
    return links


def test_create_excel_image_buffer_handles_missing_and_valid_files(tmp_path):
    assert create_excel_image_buffer(tmp_path / "missing.png") is None

    image_path = tmp_path / "valid.png"
    Image.new("RGB", (40, 30), (10, 20, 30)).save(image_path)
    result = create_excel_image_buffer(str(image_path))

    assert result["width"] == 40
    assert result["height"] == 30


def test_export_without_images_uses_no_image_cell(monkeypatch, tmp_path):
    monkeypatch.setattr(export_service, "fetch_inspection_history", lambda **kwargs: [make_record()])

    workbook_stream = export_service.export_inspections_to_excel(upload_folder=str(tmp_path))
    workbook = load_workbook(workbook_stream)
    worksheet = workbook.active

    assert worksheet.cell(row=1, column=2).value == "Images"
    assert worksheet.cell(row=2, column=2).value == "No image"
    assert len(worksheet._images) == 0


def test_export_with_six_images_keeps_images_in_one_column(monkeypatch, tmp_path):
    links = create_export_images(tmp_path, 6)
    monkeypatch.setattr(export_service, "fetch_inspection_history", lambda **kwargs: [make_record(image_links=links)])

    workbook_stream = export_service.export_inspections_to_excel(upload_folder=str(tmp_path))
    workbook = load_workbook(workbook_stream)
    worksheet = workbook.active
    headers = [worksheet.cell(row=1, column=column).value for column in range(1, worksheet.max_column + 1)]

    assert headers == ["Machine", "Images", "Issue", "Severity", "Note", "Date"]
    assert len(worksheet._images) == 6
    assert {image.anchor._from.col for image in worksheet._images} == {1}
