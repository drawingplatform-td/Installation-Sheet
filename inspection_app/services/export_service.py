from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database import to_local_time

from .inspection_service import fetch_inspection_history
from ..utils.image_utils import resolve_upload_path
from ..utils.inspection_utils import parse_image_links, severity_export_text
from ..utils.workbook_utils import add_images_to_worksheet


def export_inspections_to_excel(machine_filter="", severity_filter="", sort_order="latest", upload_folder=""):
    inspections = fetch_inspection_history(
        machine_filter=machine_filter,
        severity_filter=severity_filter,
        sort_order=sort_order,
    )

    records = []
    max_image_count = 0
    for inspection in inspections:
        image_links = parse_image_links(inspection.image_links)
        resolved_paths = [resolve_upload_path(upload_folder, image_path) for image_path in image_links]
        valid_image_paths = [image_path for image_path in resolved_paths if image_path]
        local_timestamp = to_local_time(inspection.timestamp)

        max_image_count = max(max_image_count, len(valid_image_paths))
        records.append(
            {
                "inspection": inspection,
                "valid_image_paths": valid_image_paths,
                "local_timestamp": local_timestamp,
            }
        )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inspection History"

    image_column_count = max(1, max_image_count)
    image_headers = [f"Image {index + 1}" for index in range(image_column_count)]
    headers = ["Machine"] + image_headers + ["Issue", "Severity", "Note", "Date"]
    worksheet.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    issue_column = image_column_count + 2
    severity_column = image_column_count + 3
    note_column = image_column_count + 4
    date_column = image_column_count + 5

    current_row = 2
    for record in records:
        inspection = record["inspection"]
        valid_image_paths = record["valid_image_paths"]
        local_timestamp = record["local_timestamp"]

        worksheet.append(
            [
                inspection.machine,
                *["" for _ in range(image_column_count)],
                inspection.issue or "-",
                severity_export_text(inspection.severity),
                inspection.remark or "-",
                local_timestamp.strftime("%Y-%m-%d %H:%M:%S") if local_timestamp else "-",
            ]
        )

        for column_index in [1, issue_column, severity_column, note_column, date_column]:
            worksheet.cell(row=current_row, column=column_index).alignment = Alignment(
                horizontal="center" if column_index in [severity_column, date_column] else "left",
                vertical="center",
                wrap_text=True,
            )

        for image_column in range(2, image_column_count + 2):
            worksheet.cell(row=current_row, column=image_column).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        if valid_image_paths:
            image_height = add_images_to_worksheet(
                worksheet,
                current_row,
                valid_image_paths,
                start_column_index=1,
            )
            worksheet.row_dimensions[current_row].height = max(110, image_height * 0.75)
        else:
            worksheet.cell(row=current_row, column=2).value = "No image"
            worksheet.row_dimensions[current_row].height = 110

        current_row += 1

    worksheet.column_dimensions["A"].width = 20
    for image_column in range(2, image_column_count + 2):
        worksheet.column_dimensions[get_column_letter(image_column)].width = 22
    worksheet.column_dimensions[get_column_letter(issue_column)].width = 30
    worksheet.column_dimensions[get_column_letter(severity_column)].width = 18
    worksheet.column_dimensions[get_column_letter(note_column)].width = 30
    worksheet.column_dimensions[get_column_letter(date_column)].width = 20

    excel_io = BytesIO()
    workbook.save(excel_io)
    excel_io.seek(0)
    return excel_io
